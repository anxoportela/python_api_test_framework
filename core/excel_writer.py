from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime


def _copy_worksheet(wb, source_sheet_name, new_sheet_name):
    """
    Copia una hoja de trabajo existente en el libro de trabajo y le da un nuevo nombre.

    Parámetros:
    - wb (Workbook): El objeto de libro de trabajo de `openpyxl`.
    - source_sheet_name (str): El nombre de la hoja fuente que se va a copiar.
    - new_sheet_name (str): El nuevo nombre que se le asignará a la hoja copiada.

    Retorna:
    - Worksheet: La nueva hoja de trabajo copiada, o None si la hoja fuente no existe.
    """
    if source_sheet_name not in wb.sheetnames:
        print(f"No se encontró la hoja fuente: {source_sheet_name}")
        return None
    source_sheet = wb[source_sheet_name]
    new_sheet = wb.copy_worksheet(source_sheet)
    new_sheet.title = new_sheet_name
    return new_sheet


def _create_new_sheet_with_date(wb):
    """
    Crea una nueva hoja de trabajo con un nombre basado en la fecha y hora actual.

    Parámetros:
    - wb (Workbook): El objeto de libro de trabajo de `openpyxl`.

    Retorna:
    - Worksheet: La nueva hoja de trabajo con el nombre basado en la fecha.
    """
    current_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    sheet_name = f"TestCases_{current_date}"

    # Si la hoja con ese nombre ya existe, la obtenemos, si no la creamos
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Definimos los encabezados de las columnas
    headers = [
        "TestId", "TestCase", "Status", "Error", "Method", "URL",
        "Endpoint", "ExpectedStatusCode", "ActualStatusCode",
        "Duration", "ResponseSize"
    ]
    # Insertamos los encabezados en la primera fila
    ws.append(headers)

    return ws


def _format_error_message(error_message):
    """
    Formatea el mensaje de error, agregando saltos de línea para mejorar la visualización.

    Parámetros:
    - error_message (str): El mensaje de error a formatear.

    Retorna:
    - str: El mensaje de error formateado, con saltos de línea.
    """
    if error_message:
        # Reemplazamos las llaves y las comas para mejorar la legibilidad
        formatted_message = error_message.replace("{", "\n{").replace("}", "}\n")
        return formatted_message.replace(",", "\n")
    return error_message


def _get_status_format(status, error_message):
    """
    Devuelve el formato correspondiente (estado, color y mensaje de error) en función del estado de la prueba.

    Parámetros:
    - status (str): El estado de la prueba ("PASSED", "FAILED", "SKIPPED").
    - error_message (str): El mensaje de error asociado a la prueba, si es aplicable.

    Retorna:
    - tuple: Una tupla con el estado formateado, el color y el mensaje de error (si aplica).
    """
    if status == "SKIPPED":
        return ("SKIPPED", "FFFF00", None)
    elif status == "PASSED":
        return ("PASSED", "00FF00", None)
    else:
        return ("FAILED", "FF0000", _format_error_message(error_message))


class ExcelWriter:
    """
    Clase para escribir y actualizar resultados de pruebas en un archivo Excel usando la librería `openpyxl`.
    """

    def __init__(self, file_path):
        """
        Inicializa la clase ExcelWriter con la ruta del archivo Excel.

        Parámetros:
        - file_path (str): Ruta del archivo Excel donde se escribirán los resultados.
        """
        self.file_path = file_path
        # Definimos un borde fino para aplicar a las celdas
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    def _apply_format_to_row(self, row, status, color, error_message):
        """
        Aplica formato a una fila específica de resultados en la hoja de Excel.

        Parámetros:
        - row (tuple): Fila de celdas a la que se le aplicará el formato.
        - status (str): El estado de la prueba (PASSED, FAILED, SKIPPED).
        - color (str): El color de fondo a aplicar a la celda de estado.
        - error_message (str): El mensaje de error que se mostrará si la prueba ha fallado.
        """
        # Asignamos el valor y formato a la celda de estado
        status_cell = row[13]  # Columna de estado (14ª columna)
        error_cell = row[14]  # Columna de error (15ª columna)

        status_cell.value = status
        status_cell.fill = PatternFill(start_color=color, fill_type="solid")
        status_cell.font = Font(color="000000")  # Fuente de color negro

        # Asignamos el mensaje de error y configuramos la alineación
        error_cell.value = error_message
        error_cell.alignment = Alignment(
            wrap_text=True)  # Ajusta el texto para que se muestre en varias líneas si es necesario

        # Aplicamos un borde fino a todas las celdas de la fila
        for cell in row:
            cell.border = self.thin_border

    def _update_row(self, ws, result):
        """
        Actualiza una fila de la hoja de trabajo con los resultados de la prueba.

        Parámetros:
        - ws (Worksheet): La hoja de trabajo donde se actualizarán los resultados.
        - result (dict): Diccionario con los resultados de la prueba a actualizar.
        """
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column - 1):
            if row[0].value == result["TestId"]:  # Compara el TestId para encontrar la fila correspondiente
                status, color, error_message = _get_status_format(result["Status"], result.get("Error"))
                self._apply_format_to_row(row, status, color, error_message)

    def update_results(self, results):
        """
        Actualiza los resultados de las pruebas en el archivo Excel, tanto en la hoja principal como en una nueva hoja de registro.

        Parámetros:
        - results (list): Lista de diccionarios con los resultados de las pruebas a actualizar.
        """
        try:
            if not results:
                print("No hay resultados para actualizar.")
                return

            wb = load_workbook(self.file_path)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_sheet_name = f"TestSuite_{timestamp}"

            # Comprobamos si la hoja principal 'TestSuite' existe
            if "TestSuite" in wb.sheetnames:
                main_sheet = wb["TestSuite"]
                # Actualizamos cada fila con los resultados proporcionados
                for result in results:
                    self._update_row(main_sheet, result)
            else:
                print("No se encontró la hoja principal 'TestSuite'. No se realizaron actualizaciones.")
                return

            # Copiamos la hoja 'TestSuite' y la renombramos con la marca de tiempo
            new_sheet = _copy_worksheet(wb, "TestSuite", new_sheet_name)
            if new_sheet:
                # Actualizamos los resultados en la nueva hoja
                for result in results:
                    self._update_row(new_sheet, result)

            # Guardamos el archivo Excel con las actualizaciones
            wb.save(self.file_path)
            print(f"Resultados actualizados en 'TestSuite' y en la nueva hoja '{new_sheet_name}'.")
        except Exception as e:
            print(f"Error actualizando el archivo Excel: {e}")
